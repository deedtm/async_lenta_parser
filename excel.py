from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet



class Excel:
    def __init__(self, file_name: str, sheet_name: str, columns: list[str]):
        file_name += '.xlsx'
        xls = Workbook()
        xls.remove(xls.active)
        xls.create_sheet(sheet_name)
        sheet = xls.get_sheet_by_name(sheet_name)
        sheet.append(columns)
        xls.save(file_name)

        self.xls = xls
        self.file_name = file_name
        self.columns = columns

        self.styling(sheet)

    def write(self, sheet: Worksheet, datas: list[str]):
        sheet.append(datas)
        self.xls.save(self.file_name)

    def get_sheet(self, sheet_name: str):
        return self.xls.get_sheet_by_name(sheet_name)

    def add_sheet(self, sheet_name: str):
        xls = self.xls
        xls.create_sheet(sheet_name)
        xls.save(self.file_name)

    def styling(self, sheet):
        columns_amount = len(self.columns)
        font = Font(b=True, color='000000')
        fill = PatternFill('solid', fgColor='FFFF00')

        for letter in 'ABCDEFGHIJKLMN'[:columns_amount]:
            sheet[letter + '1'].font = font
            sheet[letter + '1'].fill = fill

        self.xls.save(self.file_name)

